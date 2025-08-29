#!/usr/bin/env python3

import os
import re
import pandas as pd
import argparse
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment


def extract_cyp2d6_info(file_path):
    gene = "CYP2D6"
    file_name = os.path.basename(file_path)
    print(f"[INFO] Extracting data from: {file_name}")
    try:
        # Extract sample name and barcode from filename
        sample_name_match = re.search(r"^(.*?)_barcode(\d+)[_.]", file_name)
        if sample_name_match:
            sample_name = sample_name_match.group(1)
            barcode = sample_name_match.group(2)
        else:
            sample_name = "Unknown"
            barcode = "Unknown"
            print(f"[WARNING] Could not parse sample name or barcode from filename.")

        with open(file_path, "r") as file:
            content = file.read()

        # Try extracting population from HTML content
        group_match = re.search(r"Biogeographic Group:\s*([^\(]+)", content)
        if group_match:
            population = group_match.group(1).strip()
            print(f"[INFO] Extracted population from HTML: {population}")
        else:
            # Fallback to filename-based population extraction
            population_match = re.search(r"_population_(\w+)", file_name)
            population = population_match.group(1) if population_match else "Unknown"
            print(f"[INFO] Fallback population from filename: {population}")

        # Extract Diplotype and Phenotype
        match = re.search(
            rf"<p><font .*?><b>Gene</b>: {gene}.*?<b>Diplotype</b>: (.*?)&nbsp;.*?<b>Phenotype</b>: (.*?)</font>",
            content,
            re.DOTALL,
        )
        if match:
            diplotype = match.group(1).strip()
            phenotype = match.group(2).strip()
            phenotype = re.sub(r"\bMetabolizer\b", "", phenotype).strip()
            sample_name_with_barcode = f"{sample_name}_barcode{barcode}"
            print(
                f"[INFO] Extracted: {sample_name_with_barcode}, {population}, {diplotype} | {phenotype}"
            )
            return sample_name_with_barcode, population, f"{diplotype} | {phenotype}"
        else:
            print(f"[WARNING] No CYP2D6 match found in {file_name}")
    except Exception as e:
        print(f"[ERROR] Failed to extract info from {file_name}: {e}")
    return f"Unknown_barcode{barcode}", "Unknown", None


def main():
    parser = argparse.ArgumentParser(
        description="Extract CYP2D6 information from HTML files and save to Excel."
    )
    parser.add_argument(
        "-i",
        "--input_directory",
        required=True,
        help="Directory containing HTML files.",
    )
    parser.add_argument(
        "-o", "--output_file", required=True, help="Output Excel file path."
    )
    args = parser.parse_args()

    input_directory = args.input_directory
    output_excel_file = args.output_file

    print(f"[INFO] Scanning directory: {input_directory}")
    try:
        input_files = [
            os.path.join(input_directory, f)
            for f in os.listdir(input_directory)
            if f.endswith(".html")
        ]
        print(f"[INFO] Found {len(input_files)} HTML files.")
    except Exception as e:
        print(f"[ERROR] Could not list files in {input_directory}: {e}")
        input_files = []

    cyp2d6_data = defaultdict(lambda: defaultdict(str))

    for file_path in input_files:
        try:
            sample_name, population, diplotype_phenotype = extract_cyp2d6_info(
                file_path
            )
            if diplotype_phenotype:
                cyp2d6_data[sample_name][population] = diplotype_phenotype
        except Exception as e:
            print(f"[ERROR] Failed to process {file_path}: {e}")

    populations_present = sorted(
        set(pop for sample_data in cyp2d6_data.values() for pop in sample_data.keys())
    )
    print(f"[INFO] Populations detected: {populations_present}")

    data = []
    for sample_name, population_data in cyp2d6_data.items():
        row = [sample_name]
        for pop in populations_present:
            row.append(population_data.get(pop, ""))
        data.append(row)

    columns = ["SampleID"] + populations_present
    df = pd.DataFrame(data, columns=columns)

    citation_row = [
        "Table citation: For each population, results are Diplotype | Phenotype"
    ] + [""] * len(populations_present)
    df.loc[len(df)] = citation_row

    print(f"[INFO] Writing Excel file: {output_excel_file}")
    try:
        df.to_excel(output_excel_file, index=False)
    except Exception as e:
        print(f"[ERROR] Failed to write Excel file: {e}")
        return

    try:
        wb = load_workbook(output_excel_file)
        ws = wb.active

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        print("[INFO] Applying cell borders...")
        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.border = thin_border

        print("[INFO] Adjusting column widths...")
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception as e:
                    print(f"[WARNING] Error calculating cell length: {e}")
            ws.column_dimensions[column_letter].width = max_length + 2

        citation_cell = ws.cell(row=ws.max_row, column=1)
        citation_cell.alignment = Alignment(wrap_text=True, horizontal="left")

        wb.save(output_excel_file)
        print(f"[SUCCESS] Excel file saved with formatting: {output_excel_file}")
    except Exception as e:
        print(f"[ERROR] Failed to format Excel file: {e}")
        return


if __name__ == "__main__":
    main()
