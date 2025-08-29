#!/usr/bin/env python3

import pandas as pd
import glob
import os
import argparse
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment


def combine_pypgx_results(indir, out_prefix):
    print(f"[INFO] Searching for TSV files in: {indir}")
    all_files = glob.glob(os.path.join(indir, "*.tsv"))
    if not all_files:
        print("[WARNING] No TSV files found.")
    else:
        print(f"[INFO] Found {len(all_files)} TSV files.")

    df_list = []

    for filename in all_files:
        print(f"[INFO] Processing file: {filename}")
        try:
            df = pd.read_csv(filename, sep="\t", index_col=0)
        except Exception as e:
            print(f"[ERROR] Failed to read {filename}: {e}")
            continue

        # Determine the pipeline based on filename
        if "NGS" in filename:
            pipeline = "NGS"
        elif "TGS" in filename:
            pipeline = "TGS"
        else:
            pipeline = "Unknown"
        print(f"[INFO] Assigned pipeline: {pipeline}")

        df["Pipeline"] = pipeline
        df = df.reset_index()

        if df.columns[0] == "index":
            df = df.rename(columns={"index": "SampleID"})
            print(f"[INFO] Renamed index column to 'SampleID'")

        df_list.append(df)

    if not df_list:
        print("[ERROR] No valid dataframes to combine. Exiting.")
        return

    print("[INFO] Concatenating dataframes...")
    combined_df = pd.concat(df_list, axis=0)

    # Ensure 'SampleID' is the first column
    cols = combined_df.columns.tolist()
    if "SampleID" in cols:
        cols.insert(0, cols.pop(cols.index("SampleID")))
        combined_df = combined_df[cols]
        print("[INFO] Reordered columns to place 'SampleID' first.")

    # Reorder columns to put Pipeline after Genotype
    if "Genotype" in cols and "Pipeline" in cols:
        genotype_index = cols.index("Genotype")
        cols.insert(genotype_index + 1, cols.pop(cols.index("Pipeline")))
        combined_df = combined_df[cols]
        print("[INFO] Moved 'Pipeline' column after 'Genotype'.")

    # Sort by SampleID and Pipeline
    if "SampleID" in combined_df.columns and "Pipeline" in combined_df.columns:
        combined_df = combined_df.sort_values(by=["SampleID", "Pipeline"])
        print("[INFO] Sorted combined dataframe by SampleID and Pipeline.")

    # Write to TSV
    tsv_file = f"{out_prefix}.tsv"
    try:
        combined_df.to_csv(tsv_file, sep="\t", index=False)
        print(f"[INFO] TSV file written: {tsv_file}")
    except Exception as e:
        print(f"[ERROR] Failed to write TSV file: {e}")

    # Write to Excel
    excel_file = f"{out_prefix}.xlsx"
    try:
        combined_df.to_excel(excel_file, index=False, engine="openpyxl")
        print(f"[INFO] Excel file written: {excel_file}")
    except Exception as e:
        print(f"[ERROR] Failed to write Excel file: {e}")
        return

    # Format Excel
    try:
        wb = load_workbook(excel_file)
        ws = wb.active

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        print("[INFO] Applying cell borders...")
        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.border = thin_border

        print("[INFO] Adjusting column widths...")
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception as e:
                    print(f"[WARNING] Error calculating cell length: {e}")
            ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(excel_file)
        print(f"[INFO] Excel formatting complete. File saved: {excel_file}")
    except Exception as e:
        print(f"[ERROR] Failed to format Excel file: {e}")

    print(f"[SUCCESS] Combined results saved to {tsv_file} and {excel_file}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Combine PyPGx results from multiple TSV files into TSV and XLSX."
    )
    parser.add_argument(
        "-i", "--indir", required=True, help="Input directory containing TSV files"
    )
    parser.add_argument(
        "-o", "--outprefix", required=True, help="Output file prefix (no extension)"
    )
    args = parser.parse_args()

    combine_pypgx_results(args.indir, args.outprefix)
